import email
import random
import string
import bcrypt
from flask import Flask, render_template, request, redirect, url_for, flash, session, g
from flask_wtf.csrf import CSRFProtect
import sqlite3
from datetime import datetime, date, timedelta
import os
from contextlib import closing
import logging
from functools import wraps
from flask import jsonify
import csv
import io
from flask import Response  
from math import radians, sin, cos, sqrt, atan2
from flask import current_app
import os
from werkzeug.utils import secure_filename
from flask import send_from_directory
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask_mail import Mail, Message
from flask_mail import Mail, Message
# from app import app


hashed = bcrypt.hashpw(b'your-new-password', bcrypt.gensalt())
print(hashed.decode('utf-8'))





# Configuration
class Config:
    SECRET_KEY = os.environ.get('SECRET_KEY') or 'your-secret-key-here'
    DATABASE = 'database.db'
    SESSION_COOKIE_SECURE = os.environ.get('FLASK_ENV') == 'production'
    BYPASS_LOCATION = False
    SESSION_COOKIE_HTTPONLY = True
    PERMANENT_SESSION_LIFETIME = 3600  # 1 hour
    WTF_CSRF_TIME_LIMIT = 3600
 
    MAIL_SERVER = 'smtp.gmail.com'
    MAIL_PORT = 587
    MAIL_USE_TLS = True
    MAIL_USERNAME = 'darshanmpreddy@gmail.com'  # Your Gmail
    MAIL_PASSWORD = 'thfo gfsh zoye idjo'       # Your App Password
    MAIL_DEFAULT_SENDER = 'darshanmpreddy@gmail.com'

app = Flask(__name__)
app.config.from_object(Config)
csrf = CSRFProtect(app)
logging.basicConfig(level=logging.INFO)
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_FILE_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

mail = Mail(app)

def test_email():
    try:
        msg = Message("Test Email", 
                     sender=app.config['MAIL_DEFAULT_SENDER'],
                     recipients=["your-email@example.com"])
        msg.body = "This is a test email"
        mail.send(msg)
        print("Email sent successfully!")
    except Exception as e:
        print(f"Failed to send email: {str(e)}")

mail = Mail(app)         
def send_email(to, subject, body):
    msg = Message(subject, recipients=[to], body=body)
    mail.send(msg) 
    
def send_password_reset_email(email, new_password):
    try:
        subject = "Your Admin Password Has Been Reset"
        body = f"""Your admin password has been reset.
        
New Password: {new_password}

Please change this password immediately after logging in.
"""
        msg = Message(subject, recipients=[email], body=body)
        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"Failed to send password reset email: {str(e)}")
        return False

def allowed_file(filename, allowed_extensions):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in allowed_extensions

      
    

# Database setup
def connect_db():
    def adapt_date(d):
        return d.isoformat()
    def convert_date(s):
        return date.fromisoformat(s.decode())
    
    sqlite3.register_adapter(date, adapt_date)
    sqlite3.register_converter("DATE", convert_date)

    conn = sqlite3.connect(app.config['DATABASE'], detect_types=sqlite3.PARSE_DECLTYPES)
    conn.row_factory = sqlite3.Row
    return conn

def get_db():
    if not hasattr(g, 'sqlite_db'):
        g.sqlite_db = connect_db()
    return g.sqlite_db

# Helper functions
def calculate_duration(start_time, end_time):
    try:
        if not start_time or not end_time:
            return "-"
            
        fmt = '%H:%M:%S'
        start = datetime.strptime(start_time, fmt)
        end = datetime.strptime(end_time, fmt)
        
        if end < start:
            end += timedelta(days=1)
            
        delta = end - start
        hours, remainder = divmod(delta.seconds, 3600)
        minutes = remainder // 60
        
        if hours > 0:
            return f"{hours}h {minutes}m"
        return f"{minutes}m"
    except Exception as e:
        logging.error(f"Duration calculation error: {str(e)}")
        return "-"

# Security decorators
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            flash('Admin access required', 'warning')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated

def employee_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('employee_id'):
            flash('Please login first', 'warning')
            return redirect(url_for('employee_login'))
        return f(*args, **kwargs)
    return decorated

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


# @app.route('/')
# def home():
#     return redirect(url_for('employee_login'))

@app.errorhandler(500)
def internal_error(error):
    app.logger.error(f"Server Error: {error}")
    return render_template('500.html'), 500

# app.py - Test email route
@app.route('/test-email')
def test_email():
    try:
        msg = Message("Test Email",
                     sender=app.config['MAIL_DEFAULT_SENDER'],
                     recipients=["darshanmpreddy@gmail.com"])  # Your email
        msg.body = "This is a test email from the system"
        mail.send(msg)
        return "Email sent successfully!"
    except Exception as e:
        return f"Failed to send email: {str(e)}"

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        try:
            name = request.form['name']
            passcode = request.form['passcode']
            db = get_db()
            
            if db.execute('SELECT id FROM employees WHERE name = ?', (name,)).fetchone():
                flash('Employee already exists', 'danger')
                return redirect(url_for('register'))
            
            db.execute('INSERT INTO employees (name, passcode) VALUES (?, ?)',
                      (name, passcode))
            db.commit()
            flash('Registration successful! Please login', 'success')
            return redirect(url_for('employee_login'))
        
        except Exception as e:
            logging.error(f"Registration error: {str(e)}")
            flash('Registration failed', 'danger')
    
    return render_template('register.html')

@app.route('/employee/login', methods=['GET', 'POST'])
def employee_login():
    if session.get('employee_id'):
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        try:
            name = request.form['name']
            passcode = request.form['passcode']
            latitude = request.form.get('latitude')
            longitude = request.form.get('longitude')
            
            app.logger.info(f"Login attempt from {name} - Coordinates: {latitude}, {longitude}")

            # Location validation
            if not validate_location(latitude, longitude):
                app.logger.warning(f"Location validation failed for {name}")
                flash('Login only allowed from office location', 'danger')
                return redirect(url_for('employee_login'))

            db = get_db()
            employee = db.execute('''SELECT id, name, passcode FROM employees 
                                   WHERE name = ?''', (name,)).fetchone()
            
            if employee and employee['passcode'] == passcode:
                today = date.today()
                now = datetime.now().strftime('%H:%M:%S')
                
                # Check for existing active sessions
                existing_active = db.execute('''
                    SELECT id FROM attendance 
                    WHERE employee_id = ? 
                    AND date = ?
                    AND logout_time IS NULL
                    AND session_type IN ('work', 'break')
                    LIMIT 1''', (employee['id'], today)).fetchone()
                
                if existing_active:
                    flash('You have an active session. Please logout first', 'warning')
                    return redirect(url_for('employee_login'))

                # Check for completed work session today
                existing_work = db.execute('''
                    SELECT id FROM attendance 
                    WHERE employee_id = ? 
                    AND date = ? 
                    AND session_type = 'work'
                    ORDER BY login_time DESC LIMIT 1''',
                    (employee['id'], today)).fetchone()

                # Create new work session if no existing completed session
                if not existing_work:
                    db.execute('''INSERT INTO attendance 
                                (employee_id, date, login_time, session_type)
                                VALUES (?, ?, ?, 'work')''',
                                (employee['id'], today, now))
                
                session['employee_id'] = employee['id']
                session['employee_name'] = employee['name']
                db.commit()
                return redirect(url_for('dashboard'))
            
            flash('Invalid credentials', 'danger')
        
        except Exception as e:
            logging.error(f"Login error: {str(e)}")
            flash('Authentication failed', 'danger')
    
    return render_template('employee_login.html')

def validate_location(lat, lng):
    """Check if coordinates are within office premises using Haversine formula"""
    # Bypass location check if configured
    if app.config.get('BYPASS_LOCATION'):
        app.logger.info("Bypassing location validation")
        return True
        
    try:
        if not lat or not lng:
            app.logger.error("Location validation failed: Missing coordinates")
            return False
            
        # Convert coordinates to float
        lat = float(lat)
        lng = float(lng)
        
        # Office coordinates (UPDATE THESE WITH YOUR ACTUAL OFFICE LOCATION)
        OFFICE_LAT = 13.0946443  
        OFFICE_LNG = 77.5804383  
        ALLOWED_RADIUS = 100000 # 100 meters
        
        # Haversine formula implementation
        R = 6371000  # Earth radius in meters
        φ1 = radians(OFFICE_LAT)
        φ2 = radians(lat)
        Δφ = radians(lat - OFFICE_LAT)
        Δλ = radians(lng - OFFICE_LNG)

        a = (sin(Δφ/2) ** 2 + cos(φ1) * cos(φ2) * sin(Δλ/2) ** 2)
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        distance = R * c  # Distance in meters

        app.logger.info(f"Location check - Distance: {distance:.2f}m from office")
        return distance <= ALLOWED_RADIUS
        
    except Exception as e:
        app.logger.error(f"Location validation error: {str(e)}")
        return False

def get_employee_stats(employee_id, db):
    """Helper function to get employee statistics"""
    today_str = date.today().isoformat()
    
    # Get working days count
    working_days = db.execute('''
        SELECT COUNT(DISTINCT a.date) 
        FROM attendance a
        WHERE a.employee_id = ? 
        AND a.session_type = 'work'
        AND NOT EXISTS (
            SELECT 1 FROM holidays h 
            WHERE h.date = a.date
        )
        AND NOT EXISTS (
            SELECT 1 FROM attendance a2
            WHERE a2.employee_id = a.employee_id
            AND a2.date = a.date
            AND a2.session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
            AND a2.status = 'approved'
        )
    ''', (employee_id,)).fetchone()[0] or 0
    
    # Get total hours - FIXED QUERY
    total_hours = db.execute('''
        SELECT SUM((
            strftime('%s', last_logout) - 
            strftime('%s', first_login)
          )/3600.0
        ) FROM (
            SELECT 
                date,
                MIN(login_time) as first_login,
                MAX(logout_time) as last_logout
            FROM attendance
            WHERE employee_id = ?
            AND session_type = 'work'
            GROUP BY date
        ) AS sub
    ''', (employee_id,)).fetchone()[0] or 0

    # Get leave counts
    paid_leaves = db.execute('''
        SELECT COUNT(*) FROM attendance 
        WHERE employee_id = ? 
        AND session_type = 'paid_leave'
        AND status = 'approved'
    ''', (employee_id,)).fetchone()[0] or 0

    loss_of_pay_leaves = db.execute('''
        SELECT COUNT(*) FROM attendance 
        WHERE employee_id = ? 
        AND session_type = 'loss_of_pay'
        AND status = 'approved'
    ''', (employee_id,)).fetchone()[0] or 0

    half_day_leaves = db.execute('''
        SELECT COUNT(*) FROM attendance 
        WHERE employee_id = ? 
        AND session_type IN ('half_day', 'morning_half', 'afternoon_half')
        AND status = 'approved'
    ''', (employee_id,)).fetchone()[0] or 0
    
    week_off_leaves = db.execute('''
        SELECT COUNT(*) FROM attendance 
        WHERE employee_id = ? 
        AND session_type = 'week_off'
        AND status = 'approved'
    ''', (employee_id,)).fetchone()[0] or 0

    # Get last active time
    last_active = db.execute('''
        SELECT MAX(date || ' ' || logout_time) 
        FROM attendance 
        WHERE employee_id = ?
    ''', (employee_id,)).fetchone()[0] or 0

    # Check if currently active
    is_active = db.execute('''
        SELECT EXISTS(
            SELECT 1 FROM attendance 
            WHERE employee_id = ? 
            AND date = ? 
            AND logout_time IS NULL
            AND session_type IN ('work', 'break')
        )
    ''', (employee_id, today_str)).fetchone()[0]

    return {
        'working_days': working_days,
        'total_hours': total_hours,
        'paid_leaves': paid_leaves,
        'loss_of_pay_leaves': loss_of_pay_leaves,
        'half_day_leaves': half_day_leaves,
        'week_off_leaves': week_off_leaves,
        'last_active': last_active,
        'is_active': is_active
    }

@app.route('/dashboard')
@employee_required
def dashboard():
    try:
        db = get_db()
        today = date.today()
        
        today_records = db.execute('''
            SELECT date, login_time, logout_time, session_type
            FROM attendance 
            WHERE employee_id = ? AND date = ?
            ORDER BY login_time DESC
        ''', (session['employee_id'], today)).fetchall()

        # Get employee stats
        stats = get_employee_stats(session['employee_id'], db)

        return render_template('dashboard.html',
                            name=session['employee_name'],
                            today_records=today_records,
                            stats=stats,
                            calculate_duration=calculate_duration)
    except Exception as e:
        logging.error(f"Dashboard error: {str(e)}")
        flash('Failed to load records', 'danger')
        return redirect(url_for('employee_login'))

@app.route('/break/start')
@employee_required
def start_break():
    try:
        today = date.today()
        now = datetime.now().strftime('%H:%M:%S')
        db = get_db()
        
        # Check if there's an active work session
        active_work = db.execute('''
            SELECT id FROM attendance 
            WHERE employee_id = ? 
            AND date = ? 
            AND session_type = 'work'
            AND logout_time IS NULL
            LIMIT 1
        ''', (session['employee_id'], today)).fetchone()
        
        if not active_work:
            flash('You must be in a work session to start a break', 'warning')
            return redirect(url_for('dashboard'))
        
        # End current work session
        db.execute('UPDATE attendance SET logout_time = ? WHERE id = ?',
                  (now, active_work['id']))
        
        # Start break session
        db.execute('''
            INSERT INTO attendance 
            (employee_id, date, login_time, session_type)
            VALUES (?, ?, ?, 'break')
        ''', (session['employee_id'], today, now))
        
        db.commit()
        flash('Break started successfully', 'success')
    except Exception as e:
        logging.error(f"Break start error: {str(e)}")
        flash('Failed to start break', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/break/end')
@employee_required
def end_break():
    try:
        today = date.today()
        now = datetime.now().strftime('%H:%M:%S')
        db = get_db()
        
        # Check if there's an active break session
        active_break = db.execute('''
            SELECT id FROM attendance 
            WHERE employee_id = ? 
            AND date = ? 
            AND session_type = 'break'
            AND logout_time IS NULL
            LIMIT 1
        ''', (session['employee_id'], today)).fetchone()
        
        if not active_break:
            flash('No active break session to end', 'warning')
            return redirect(url_for('dashboard'))
        
        # End current break session
        db.execute('UPDATE attendance SET logout_time = ? WHERE id = ?',
                  (now, active_break['id']))
        
        # Start new work session
        db.execute('''
            INSERT INTO attendance 
            (employee_id, date, login_time, session_type)
            VALUES (?, ?, ?, 'work')
        ''', (session['employee_id'], today, now))
        
        db.commit()
        flash('Break ended successfully', 'success')
    except Exception as e:
        logging.error(f"Break end error: {str(e)}")
        flash('Failed to end break', 'danger')
    
    return redirect(url_for('dashboard'))

# app.py
# app.py
@app.route('/logout')
@employee_required
def logout():
    try:
        today = date.today()
        now = datetime.now().strftime('%H:%M:%S')
        db = get_db()
        
        # Check if logout already exists for today
        existing_logout = db.execute('''
            SELECT 1 FROM attendance 
            WHERE employee_id = ? 
            AND date = ?
            AND session_type = 'logout'
            LIMIT 1
        ''', (session['employee_id'], today)).fetchone()
        
        if existing_logout:
            flash('You have already logged out today', 'warning')
            session.clear()
            return redirect(url_for('employee_login'))
        
        # Get first login time
        first_login = db.execute('''
            SELECT MIN(login_time) as first_login 
            FROM attendance 
            WHERE employee_id = ? 
            AND date = ?
            AND session_type = 'work'
        ''', (session['employee_id'], today)).fetchone()['first_login']
        
        # Update all open sessions
        db.execute('''
            UPDATE attendance SET logout_time = ?
            WHERE employee_id = ? 
            AND date = ?
            AND logout_time IS NULL
        ''', (now, session['employee_id'], today))
        
        # Create logout session record only if first_login exists
        if first_login:
            db.execute('''
                INSERT INTO attendance 
                (employee_id, date, login_time, logout_time, session_type)
                VALUES (?, ?, ?, ?, 'logout')
            ''', (session['employee_id'], today, first_login, now))
        
        db.commit()
        session.clear()
        flash('Successfully logged out for today', 'success')
    except Exception as e:
        logging.error(f"Logout error: {str(e)}")
        flash('Logout failed', 'danger')

    return redirect(url_for('employee_login'))

# Employee calendar
@app.route('/api/attendance')
@employee_required
def attendance_data():
    try:
        db = get_db()
        records = db.execute('''
            SELECT date, session_type, status
            FROM attendance 
            WHERE employee_id = ?
            AND session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
            AND status IN ('approved', 'pending')  
            ORDER BY date DESC
        ''', (session['employee_id'],)).fetchall()
        
        events = []
        for record in records:
            # Check if this date is a holiday
            is_holiday = db.execute('SELECT 1 FROM holidays WHERE date = ?', 
                                  (record['date'],)).fetchone()
            
            # Skip if it's a holiday and the leave was approved (meaning it was cancelled)
            if is_holiday and record['status'] == 'approved':
                continue
                
            events.append({
                'title': f"{record['session_type'].replace('_', ' ').title()}",
                'start': str(record['date']),
                'extendedProps': {
                    'session_type': record['session_type'],
                    'status': record['status']
                },
                'color': '#28a745' if record['session_type'] == 'paid_leave' else
                         '#dc3545' if record['session_type'] == 'loss_of_pay' else
                         '#ffc107' if record['session_type'] in ['half_day', 'morning_half', 'afternoon_half'] else
                         '#17a2b8'
            })
        return jsonify(events)
    except Exception as e:
        return jsonify([])

# In your admin_login route in app.py
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'GET':
        if session.get('admin_logged_in'):
            return redirect(url_for('admin'))
        return render_template('admin_login.html')
    
    if request.method == 'POST':
        try:
            username = request.form['username']
            password = request.form['password'].encode('utf-8')
            db = get_db()
            admin = db.execute('''SELECT username, password FROM admins 
                                WHERE username = ?''', (username,)).fetchone()
            
            if admin and bcrypt.checkpw(password, admin['password'].encode('utf-8')):
                session.permanent = True
                session['admin_logged_in'] = True
                session['admin_username'] = admin['username']
                flash('Login successful', 'success')
                return redirect(url_for('admin'))
            
            flash('Invalid credentials', 'danger')
            return render_template('admin_login.html')
        
        except Exception as e:
            logging.error(f"Admin login error: {str(e)}")
            flash('Authentication failed', 'danger')
            return render_template('admin_login.html')

@app.route('/admin/logout')
@admin_required
def admin_logout():
    session.clear()
    flash('Successfully logged out', 'success')
    return redirect(url_for('admin_login'))

@app.route('/admin')
@admin_required
def admin():
    try:
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login'))
            
        db = get_db()
        today = date.today()
        today_str = today.isoformat()  
        
        employees = db.execute('''
    SELECT 
        e.id, 
        e.employee_id,
        e.name,
        e.department,
        (
            SELECT COUNT(DISTINCT a.date) 
            FROM attendance a
            WHERE a.employee_id = e.id 
            AND a.session_type = 'work'
            AND NOT EXISTS (
                SELECT 1 FROM holidays h 
                WHERE h.date = a.date
            )
            AND NOT EXISTS (
                SELECT 1 FROM attendance a2
                WHERE a2.employee_id = a.employee_id
                AND a2.date = a.date
                AND a2.session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
                AND a2.status = 'approved'
            )
        ) as working_days,
        MAX(a.date || ' ' || a.logout_time) as last_active,
        EXISTS(
            SELECT 1 FROM attendance 
            WHERE employee_id = e.id 
            AND date = ?
            AND logout_time IS NULL
            AND session_type IN ('work', 'break')
        ) as is_active
    FROM employees e
    LEFT JOIN attendance a ON e.id = a.employee_id
    GROUP BY e.id
    ORDER BY e.id ASC  
''', (today_str,)).fetchall()
        
            
        employees_count = db.execute('SELECT COUNT(*) FROM employees').fetchone()[0]
        
        active_today_count = db.execute('''
            SELECT COUNT(DISTINCT employee_id) 
            FROM attendance 
            WHERE date = ? 
            AND session_type = 'work'
            AND logout_time IS NULL
        ''', (today_str,)).fetchone()[0]  # Use today_str here

        on_break_count = db.execute('''
            SELECT COUNT(DISTINCT employee_id) 
            FROM attendance 
            WHERE date = ? 
            AND session_type = 'break'
            AND logout_time IS NULL
        ''', (today_str,)).fetchone()[0]  # Use today_str here
        
        recent_activity = db.execute('''
            SELECT a.date, e.name as employee_name, 
                   a.login_time, a.logout_time, a.session_type
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            ORDER BY a.date DESC, a.login_time DESC
            LIMIT 50
        ''').fetchall()
        
        return render_template('admin.html',
                            employees=employees,
                            employees_count=employees_count,
                            active_today_count=active_today_count,
                            on_break_count=on_break_count,
                            recent_activity=recent_activity)
    except Exception as e:
        logging.error(f"Admin dashboard error: {str(e)}")
        flash('Failed to load dashboard data', 'danger')
        return redirect(url_for('admin_login'))

# Admin calendar
@app.route('/api/admin/attendance')
@admin_required
def admin_attendance_api():
    try:
        db = get_db()
        records = db.execute('''
            SELECT a.date, e.name, a.session_type, a.status
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            WHERE a.session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
            AND status IN ('approved', 'pending', 'rejected')
            ORDER BY a.date DESC
        ''').fetchall()
        
        events = []
        for record in records:
            # Check if this date is a holiday
            is_holiday = db.execute('SELECT 1 FROM holidays WHERE date = ?', 
                                  (record['date'],)).fetchone()
            
            # Skip if it's a holiday and the leave was approved (meaning it was cancelled)
            if is_holiday and record['status'] == 'approved':
                continue
                
            events.append({
                'title': f"{record['name']} - {record['session_type'].replace('_', ' ').title()}",
                'start': str(record['date']),
                'allDay': True,
                'extendedProps': {
                    'session_type': record['session_type'],
                    'status': record['status']
                },
                'color': '#28a745' if record['session_type'] == 'paid_leave' else
                         '#dc3545' if record['session_type'] == 'loss_of_pay' else
                         '#ffc107' if record['session_type'] in ['half_day', 'morning_half', 'afternoon_half'] else
                         '#17a2b8'
            })
        return jsonify(events)
    except Exception as e:
        return jsonify([])

@app.route('/admin/employees')
@admin_required
def admin_employees():
    try:
        db = get_db()
        employees = db.execute('''
            SELECT e.id, e.name, 
            MAX(a.date) as last_active,
            COUNT(DISTINCT a.date) as attendance_days,
            EXISTS(
                SELECT 1 FROM attendance 
                WHERE employee_id = e.id 
                AND date = date('now') 
                AND logout_time IS NULL
            ) as is_active
            FROM employees e
            LEFT JOIN attendance a ON e.id = a.employee_id
            GROUP BY e.id
            ORDER BY e.name
        ''').fetchall()
        
        if not employees:
            flash('No employees found', 'info')
            return redirect(url_for('admin'))
            
        return render_template('admin_employees.html', employees=employees)
    except Exception as e:
        logging.error(f"Admin employees error: {str(e)}")
        flash('Failed to load employee records', 'danger')
        return redirect(url_for('admin'))

@app.route('/admin/employee/<int:id>/daily/<date_str>')
@admin_required
def admin_employee_daily(id, date_str):
    try:
        # Parse date safely
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        db = get_db()
        
        # Fetch records using explicit date formatting
        records = db.execute('''
    SELECT id, login_time, logout_time, session_type
    FROM attendance 
    WHERE employee_id = ? 
    AND date = ?
    AND (
        session_type IN ('work', 'break') OR
        (session_type = 'logout' AND id = (
            SELECT MIN(id) FROM attendance 
            WHERE employee_id = ? 
            AND date = ? 
            AND session_type = 'logout'
            LIMIT 1
        ))
    )
    ORDER BY login_time ASC
''', (id, selected_date.isoformat, id, selected_date)).fetchall()

        return render_template('admin_employee_daily.html',
                           records=records,
                           date=selected_date,
                           calculate_duration=calculate_duration,
                           employee={'id': id, 'name': 'Employee'})  # Add required fields

    except ValueError as e:
        logging.error(f"Date parsing error: {str(e)}")
        flash('Invalid date format', 'danger')
        return redirect(url_for('admin_employee_detail', id=id))
    except Exception as e:
        logging.error(f"Database error: {str(e)}")
        flash('Error loading records', 'danger')
        return redirect(url_for('admin_employee_detail', id=id))
    except Exception as e:
        logging.error(f"Error in admin_employee_daily: {str(e)}", exc_info=True)
        flash('Error loading daily view', 'danger')
        return redirect(url_for('admin_employee_detail', id=id))
        
# Edit Employee
@app.route('/admin/employee/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_employee(id):
    db = get_db()
    employee = db.execute('SELECT * FROM employees WHERE id = ?', (id,)).fetchone()
    
    if request.method == 'POST':
        name = request.form['name']
        department = request.form['department']
        db.execute('UPDATE employees SET name = ?, department = ? WHERE id = ?',
                  (name, department, id))
        db.commit()
        flash('Employee updated successfully', 'success')
        return redirect(url_for('admin_employee_detail', id=id))
    
    return render_template('edit_employee.html', employee=employee)

    


# Reset Password
@app.route('/admin/employee/<int:id>/reset_password', methods=['GET', 'POST'])
@admin_required
def reset_password(id):
    db = get_db()
    if request.method == 'POST':
        new_password = request.form['new_password']
        db.execute('UPDATE employees SET passcode = ? WHERE id = ?',
                  (new_password, id))
        db.commit()
        flash('Password reset successfully', 'success')
        return redirect(url_for('admin_employee_detail', id=id))
    
    return render_template('reset_password.html', employee_id=id)

# Deactivate Employee
@app.route('/admin/employee/<int:id>/toggle_status')
@admin_required
def toggle_employee_status(id):
    db = get_db()
    employee = db.execute('SELECT * FROM employees WHERE id = ?', (id,)).fetchone()
    new_status = not employee['is_active']
    
    db.execute('UPDATE employees SET is_active = ? WHERE id = ?',
              (new_status, id))
    db.commit()
    
    flash(f'Employee {"activated" if new_status else "deactivated"} successfully', 'success')
    return redirect(url_for('admin_employee_detail', id=id))

# Delete Employee
@app.route('/admin/employee/<int:id>/delete')
@admin_required
def delete_employee(id):
    try:
        db = get_db()
        db.execute('DELETE FROM employees WHERE id = ?', (id,))
        db.execute('DELETE FROM attendance WHERE employee_id = ?', (id,))
        db.commit()
        flash('Employee deleted successfully', 'success')
    except Exception as e:
        logging.error(f"Delete error: {str(e)}")
        flash('Deletion failed', 'danger')
    
    return redirect(url_for('admin'))

@app.route('/dashboard/<date_str>')
@employee_required
def daily_view(date_str):
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        db = get_db()
        
        records = db.execute('''
            SELECT id, login_time, logout_time, session_type
            FROM attendance 
            WHERE employee_id = ? AND date = ?
            ORDER BY login_time ASC
        ''', (session['employee_id'], selected_date)).fetchall()
        
        return render_template('daily_view.html',
                           name=session['employee_name'],
                           date=selected_date,
                           records=records,
                           calculate_duration=calculate_duration)
    except Exception as e:
        flash('Invalid date format', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/admin/calendar')
@admin_required
def admin_calendar():
    return render_template('admin_calendar.html')

@app.route('/admin/daily/<date_str>')
@admin_required
def admin_daily_view(date_str):
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        db = get_db()
        
        records = db.execute('''
            SELECT a.id, e.name, a.login_time, a.logout_time, a.session_type
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            WHERE a.date = ?
            ORDER BY e.name, a.login_time ASC
        ''', (selected_date,)).fetchall()
        
        return render_template('admin_daily_view.html',
                           date=selected_date,
                           records=records,
                           calculate_duration=calculate_duration)
    except Exception as e:
        flash('Invalid date format', 'danger')
        return redirect(url_for('admin_calendar'))

# Add new route for adding employees
@app.route('/admin/add_employee', methods=['GET', 'POST'])
@admin_required
def add_employee():
    if request.method == 'POST':
        try:
            employee_id = request.form['employee_id']  # New field
            name = request.form['name']
            department = request.form['department']
            passcode = request.form['passcode']
            db = get_db()
            
            # Check if employee ID already exists
            if db.execute('SELECT id FROM employees WHERE employee_id = ?', (employee_id,)).fetchone():
                flash('Employee ID already exists', 'danger')
                return redirect(url_for('add_employee'))
            
            db.execute('''INSERT INTO employees 
                        (employee_id, name, department, passcode) 
                        VALUES (?, ?, ?, ?)''',
                      (employee_id, name, department, passcode))
            db.commit()
            flash('Employee added successfully', 'success')
            return redirect(url_for('admin'))
        
        except Exception as e:
            logging.error(f"Add employee error: {str(e)}")
            flash('Failed to add employee', 'danger')
    
    return render_template('add_employee.html')

# @app.route('/admin/employee/<string:employee_id>')
@app.route('/admin/employee/<int:id>')
@admin_required
def admin_employee_detail(id):
    try:
        db = get_db()
        employee = db.execute('SELECT * FROM employees WHERE id = ?', (id,)).fetchone()
        today_str = date.today().isoformat()
        
        # Get working days count
        # Update the working_days calculation in the admin_employee_detail route
        working_days = db.execute('''
    SELECT COUNT(DISTINCT a.date) 
    FROM attendance a
    WHERE a.employee_id = ? 
    AND a.session_type = 'work'
    AND NOT EXISTS (
        SELECT 1 FROM holidays h 
        WHERE h.date = a.date
    )
    AND NOT EXISTS (
        SELECT 1 FROM attendance a2
        WHERE a2.employee_id = a.employee_id
        AND a2.date = a.date
        AND a2.session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
        AND a2.status = 'approved'
    )
''', (id,)).fetchone()[0] or 0
        
        # Get total hours (fixed query)
        total_hours = db.execute('''
            SELECT SUM((
                strftime('%s', last_logout) - 
                strftime('%s', first_login)
              ))/3600.0
            FROM (
                SELECT 
                    date,
                    MIN(login_time) as first_login,
                    MAX(logout_time) as last_logout
                FROM attendance
                WHERE employee_id = ?
                AND session_type = 'work'
                GROUP BY date
            ) AS sub
        ''', (id,)).fetchone()[0] or 0

        # Get leave counts
        paid_leaves = db.execute('''
            SELECT COUNT(*) FROM attendance 
            WHERE employee_id = ? 
            AND session_type = 'paid_leave'
            AND status = 'approved'
        ''', (id,)).fetchone()[0]

        loss_of_pay_leaves = db.execute('''
            SELECT COUNT(*) FROM attendance 
            WHERE employee_id = ? 
            AND session_type = 'loss_of_pay'
            AND status = 'approved'
        ''', (id,)).fetchone()[0]

        half_day_leaves = db.execute('''
            SELECT COUNT(*) FROM attendance 
            WHERE employee_id = ? 
            AND session_type IN ('half_day', 'morning_half', 'afternoon_half')
            AND status = 'approved'
        ''', (id,)).fetchone()[0] or 0
        
        week_off_leaves = db.execute('''
            SELECT COUNT(*) FROM attendance 
            WHERE employee_id = ? 
            AND session_type = 'week_off'
            AND status = 'approved'
        ''', (id,)).fetchone()[0] or 0

        # Get last active time
        last_active = db.execute('''
            SELECT MAX(date || ' ' || logout_time) 
            FROM attendance 
            WHERE employee_id = ?
        ''', (id,)).fetchone()[0] or 0

        # Check if currently active - FIXED QUERY
        is_active = db.execute('''
            SELECT EXISTS(
                SELECT 1 FROM attendance 
                WHERE employee_id = ? 
                AND date = ? 
                AND logout_time IS NULL
                AND session_type IN ('work', 'break')
            )
        ''', (id, today_str)).fetchone()[0]

        return render_template('admin_employee_detail.html',
                            employee=employee,
                            stats={
                                'working_days': working_days,
                                'total_hours': total_hours or 0,
                                'paid_leaves': paid_leaves,
                                'loss_of_pay_leaves': loss_of_pay_leaves,
                                'half_day_leaves': half_day_leaves,
                                'week_off_leaves': week_off_leaves, 
                                'is_active': is_active,
                                'last_active': last_active
                            },
                            records=db.execute('''
                                SELECT * FROM attendance 
                                WHERE employee_id = ? 
                                ORDER BY date DESC 
                                LIMIT 30
                            ''', (id,)).fetchall(),
                            calculate_duration=calculate_duration)
    except Exception as e:
        logging.error(f"Employee detail error: {str(e)}")
        flash('Failed to load employee details', 'danger')
        return redirect(url_for('admin'))

# @app.route('/admin/employee/<int:id>/daily/<date_str>')
# @admin_required
# def admin_employee_daily(id, date_str):
#     try:
#         selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
#         db = get_db()
        
#         records = db.execute('''
#             SELECT id, login_time, logout_time, session_type
#             FROM attendance 
#             WHERE employee_id = ? AND date = ?
#             ORDER BY login_time ASC
#         ''', (id, selected_date)).fetchall()
        
#         return render_template('admin_employee_daily.html',
#                            records=records,
#                            date=selected_date,
#                            calculate_duration=calculate_duration)
#     except Exception as e:
#         flash('Invalid date format', 'danger')
#         return redirect(url_for('admin_employee_detail', id=id))



# Add route for exporting attendance
@app.route('/admin/export_attendance/<int:month>')
@admin_required
def export_attendance(month):
    try:
        # Get current year
        year = datetime.now().year
        
        # Validate month
        if month < 1 or month > 12:
            flash('Invalid month selected', 'danger')
            return redirect(url_for('admin'))
        
        # Get number of days in the requested month
        import calendar
        _, total_days = calendar.monthrange(year, month)
        month_name = calendar.month_name[month]
        
        # Get all employees
        db = get_db()
        employees = db.execute('''
            SELECT e.id, e.employee_id, e.name, e.department 
            FROM employees e
        ''').fetchall()
        
        # Get holidays in the month
        holidays = db.execute('''
            SELECT date FROM holidays 
            WHERE strftime('%m', date) = ? AND strftime('%Y', date) = ?
        ''', (f"{month:02d}", str(year))).fetchall()
        holiday_dates = [h['date'] for h in holidays]
        holiday_count = len(holiday_dates)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"
        
        # Write header
        ws.append(["Monthly Attendance Summary"])
        ws.append(["Month", f"{month_name} {year}"])
        ws.append(["Total Days", total_days])
        ws.append(["Holidays", holiday_count])
        ws.append([])  # Empty row
        
        # Write column headers - Swapped "Employee Working Days" and "Total Working Days"
        headers = [
            "Employee ID",
            "Employee Name",
            "Department",
            "Paid Leaves",
            "Week-Off Leaves",
            "Half Days (Count as 0.5)",
            "Loss of Pay Leaves",
            "Holidays",
            "Employee Working Days (Present)",  # Moved before Total Working Days
            "Total Working Days",  # Moved after Employee Working Days
            "Total Days in Month",
            "Total Payable Days"
        ]
        ws.append(headers)
        
        # Define color fills
        paid_leave_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
        week_off_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')    # Light blue
        half_day_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')    # Light yellow
        loss_of_pay_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # Light red
        holidays_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')    # Light orange
        working_days_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Pale green
        days_present_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid') # Pale blue
        total_days_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # Pale yellow
        payable_days_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Pale green (same as working days)
        
        # Write data for each employee
        for employee in employees:
            # Get paid leaves count
            paid_leaves = db.execute('''
                SELECT COUNT(*) FROM attendance 
                WHERE employee_id = ? 
                AND session_type = 'paid_leave'
                AND status = 'approved'
                AND strftime('%m', date) = ? 
                AND strftime('%Y', date) = ?
            ''', (employee['id'], f"{month:02d}", str(year))).fetchone()[0]
            
            # Get week-off leaves count
            week_off_leaves = db.execute('''
                SELECT COUNT(*) FROM attendance 
                WHERE employee_id = ? 
                AND session_type = 'week_off'
                AND status = 'approved'
                AND strftime('%m', date) = ? 
                AND strftime('%Y', date) = ?
            ''', (employee['id'], f"{month:02d}", str(year))).fetchone()[0]
            
            # Get half leaves count - now counting as 0.5 days each
            half_day_count = db.execute('''
                SELECT COUNT(*) FROM attendance 
                WHERE employee_id = ? 
                AND session_type IN ('morning_half', 'afternoon_half')
                AND status = 'approved'
                AND strftime('%m', date) = ? 
                AND strftime('%Y', date) = ?
            ''', (employee['id'], f"{month:02d}", str(year))).fetchone()[0]
            half_day_days = half_day_count * 0.5  # Convert count to days
            
            # Get loss of pay leaves count
            loss_of_pay = db.execute('''
                SELECT COUNT(*) FROM attendance 
                WHERE employee_id = ? 
                AND session_type = 'loss_of_pay'
                AND status = 'approved'
                AND strftime('%m', date) = ? 
                AND strftime('%Y', date) = ?
            ''', (employee['id'], f"{month:02d}", str(year))).fetchone()[0]
            
            # Get working days count (matches admin dashboard logic)
            working_days = db.execute('''
                SELECT COUNT(DISTINCT a.date) 
                FROM attendance a
                WHERE a.employee_id = ? 
                AND a.session_type = 'work'
                AND strftime('%m', a.date) = ? 
                AND strftime('%Y', a.date) = ?
                AND NOT EXISTS (
                    SELECT 1 FROM holidays h 
                    WHERE h.date = a.date
                )
                AND NOT EXISTS (
                    SELECT 1 FROM attendance a2
                    WHERE a2.employee_id = a.employee_id
                    AND a2.date = a.date
                    AND a2.session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
                    AND a2.status = 'approved'
                )
            ''', (employee['id'], f"{month:02d}", str(year))).fetchone()[0] or 0
            
            # Calculate total working days (total days - Sundays - holidays - all leaves)
            total_working_days = total_days - holiday_count - week_off_leaves
            
            # Calculate Total Payable Days = Total Days - Loss of Pay - Half Days
            total_payable_days = total_days - loss_of_pay - half_day_days
            
            # Updated row with swapped columns
            row = [
                employee['employee_id'],
                employee['name'],
                employee['department'] or 'N/A',
                paid_leaves,
                week_off_leaves,
                half_day_days,  # Now showing as 0.5 days per half day
                loss_of_pay,
                holiday_count,
                working_days,  # Now in column 9 (previously column 10)
                total_working_days,  # Now in column 10 (previously column 9)
                total_days,
                total_payable_days
            ]
            ws.append(row)
            
            # Apply colors to the current row (updated column numbers for swapped fields)
            row_idx = ws.max_row
            ws.cell(row=row_idx, column=4).fill = paid_leave_fill  # Paid Leaves
            ws.cell(row=row_idx, column=5).fill = week_off_fill    # Week-Off Leaves
            ws.cell(row=row_idx, column=6).fill = half_day_fill   # Half Days Leaves
            ws.cell(row=row_idx, column=7).fill = loss_of_pay_fill  # Loss of Pay Leaves
            ws.cell(row=row_idx, column=8).fill = holidays_fill    # Holidays
            ws.cell(row=row_idx, column=9).fill = days_present_fill  # Employee Working Days (Present)
            ws.cell(row=row_idx, column=10).fill = working_days_fill  # Total Working Days
            ws.cell(row=row_idx, column=11).fill = total_days_fill   # Total Days
            ws.cell(row=row_idx, column=12).fill = payable_days_fill  # Total Payable Days
        
        # Make header row bold
        for cell in ws[6]:  # Header row is row 6
            cell.font = Font(bold=True)
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
        
        # Save workbook to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response with Excel file
        response = Response(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 
                    f"attachment;filename=attendance_{month_name}_{year}.xlsx"
            }
        )
        
        return response
        
    except Exception as e:
        logging.error(f"Export error: {str(e)}")
        flash('Failed to export attendance summary', 'danger')
        return redirect(url_for('admin'))

# Leave Management Routes

@app.route('/request_leave', methods=['GET', 'POST'])
@employee_required
def request_leave():
    if request.method == 'POST':
        try:
            leave_type = request.form['leave_type']
            leave_date_str = request.form['leave_date']
            
            if not leave_date_str:
                flash('Please select a date', 'danger')
                return redirect(url_for('request_leave'))
            
            leave_date = datetime.strptime(leave_date_str, '%Y-%m-%d').date()
            db = get_db()
            
            # Check if date is a holiday
            is_holiday = db.execute('SELECT 1 FROM holidays WHERE date = ?', 
                                   (leave_date,)).fetchone()
            if is_holiday:
                flash('Cannot request leave on a holiday', 'danger')
                return redirect(url_for('request_leave'))
            
            # Check for existing leave on same date (including all types)
            existing_leave = db.execute('''
                SELECT 1 FROM attendance 
                WHERE employee_id = ? 
                AND date = ?
                AND session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
            ''', (session['employee_id'], leave_date)).fetchone()
            
            if existing_leave:
                flash('Only one leave per day allowed', 'danger')
                return redirect(url_for('request_leave'))
            
            month = leave_date.month
            year = leave_date.year
            
            # Check paid leave limit (1 per month)
            if leave_type == 'paid_leave':
                paid_leave_count = db.execute('''
                    SELECT COUNT(*) FROM attendance
                    WHERE employee_id = ?
                    AND session_type = 'paid_leave'
                    AND strftime('%m', date) = ?
                    AND strftime('%Y', date) = ?
                    AND status = 'approved'
                ''', (session['employee_id'], f"{month:02d}", str(year))).fetchone()[0]
                
                if paid_leave_count >= 1:
                    flash('Monthly paid leave limit reached (max 1 per month)', 'danger')
                    return redirect(url_for('request_leave'))
            
            # Check week-off limit (1 per week)
            if leave_type == 'week_off':
                # Calculate week boundaries (Monday to Sunday)
                week_start = leave_date - timedelta(days=leave_date.weekday())
                week_end = week_start + timedelta(days=6)
                
                # Check for existing week-off in same week
                existing_week_off = db.execute('''
                    SELECT COUNT(*) FROM attendance 
                    WHERE employee_id = ? 
                    AND date BETWEEN ? AND ?
                    AND session_type = 'week_off'
                    AND status = 'approved'
                ''', (session['employee_id'], week_start, week_end)).fetchone()[0]
                
                if existing_week_off >= 1:
                    flash('Only one week-off allowed per week (Mon-Sun)', 'danger')
                    return redirect(url_for('request_leave'))
            
            # Insert leave request
            db_session_type = 'week_off' if leave_type == 'week_off' else leave_type
                
            db.execute('''
                INSERT INTO attendance 
                (employee_id, date, session_type, status)
                VALUES (?, ?, ?, ?)
            ''', (
                session['employee_id'], 
                leave_date, 
                db_session_type, 
                'pending'
            ))
            
            db.commit()
            flash('Leave request submitted for approval', 'success')
            return redirect(url_for('request_leave'))
            
        except ValueError as e:
            logging.error(f"Invalid date format: {str(e)}")
            flash('Invalid date format. Please select a valid date.', 'danger')
            return redirect(url_for('request_leave'))
        except Exception as e:
            logging.error(f"Leave request error: {str(e)}", exc_info=True)
            flash('Failed to submit leave request', 'danger')
    
    # For GET request - show leave request form
    db = get_db()
    holidays = db.execute('SELECT date FROM holidays').fetchall()
    holiday_dates = [h['date'] for h in holidays]
    
    return render_template('request_leave.html', 
                         holiday_dates=holiday_dates,
                         current_date=date.today().isoformat())

@app.route('/admin/leaves')
@admin_required
def pending_leaves():
    try:
        db = get_db()
        
        employee_filter = request.args.get('employee', '')
        status_filter = request.args.get('status', '')
        leave_type_filter = request.args.get('leave_type', '')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        view = request.args.get('view', 'pending')
        history_month = request.args.get('month', date.today().month)
        history_year = request.args.get('year', date.today().year)
        page = request.args.get('page', 1, type=int)
        
        today = date.today()
        current_year = today.year
        current_month = today.month
        current_month_name = today.strftime('%B')
        
        base_query = '''
            SELECT a.id, e.name, e.employee_id, a.date, a.session_type, a.status 
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            WHERE a.session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'morning_half', 'afternoon_half', 'week_off')
        '''
        
        params = []
        
        if employee_filter:
            base_query += " AND (e.name LIKE ? OR e.employee_id LIKE ?)"
            params.extend([f'%{employee_filter}%', f'%{employee_filter}%'])
        
        if status_filter:
            base_query += " AND a.status = ?"
            params.append(status_filter)
        
        if leave_type_filter:
            base_query += " AND a.session_type = ?"
            params.append(leave_type_filter)
        
        if start_date:
            base_query += " AND a.date >= ?"
            params.append(start_date)
        
        if end_date:
            base_query += " AND a.date <= ?"
            params.append(end_date)
        
        base_query += " ORDER BY a.date DESC"
        
        leaves = db.execute(base_query, params).fetchall()
        
        history_stats = {
            'paid_leaves': 0,
            'loss_of_pay': 0,
            'half_days': 0,
            'week_offs': 0,
            'total_leaves': 0
        }
        
        history_query = '''
            SELECT a.id, e.name, e.employee_id, a.date, a.session_type, a.status 
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            WHERE a.session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'morning_half', 'afternoon_half', 'week_off')
            AND strftime('%m', a.date) = ? 
            AND strftime('%Y', a.date) = ?
        '''
        
        history_params = [f"{int(history_month):02d}", str(history_year)]
        
        if employee_filter:
            history_query += " AND (e.name LIKE ? OR e.employee_id LIKE ?)"
            history_params.extend([f'%{employee_filter}%', f'%{employee_filter}%'])
        
        if leave_type_filter:
            history_query += " AND a.session_type = ?"
            history_params.append(leave_type_filter)
        
        if status_filter:
            history_query += " AND a.status = ?"
            history_params.append(status_filter)
        
        history_leaves = db.execute(history_query, history_params).fetchall()
        
        for leave in history_leaves:
            if leave['session_type'] == 'paid_leave':
                history_stats['paid_leaves'] += 1
            elif leave['session_type'] == 'loss_of_pay':
                history_stats['loss_of_pay'] += 1
            elif leave['session_type'] in ['half_day', 'morning_half', 'afternoon_half']:
                history_stats['half_days'] += 1
            elif leave['session_type'] == 'week_off':
                history_stats['week_offs'] += 1
        
        history_stats['total_leaves'] = len(history_leaves)
        
        return render_template('admin_leaves.html', 
                           leaves=leaves,
                           history_leaves=history_leaves,
                           current_year=current_year,
                           current_month=current_month,
                           current_month_name=current_month_name,
                           stats=history_stats,
                           history_pages=1,
                           current_page=page,
                           view=view)
    except Exception as e:
        logging.error(f"Leave management error: {str(e)}")
        flash('Failed to load leave requests', 'danger')
        return redirect(url_for('admin'))

@app.route('/admin/leave/<int:id>/<action>')
@admin_required
def handle_leave(id, action):
    db = get_db()
    try:
        leave = db.execute('SELECT * FROM attendance WHERE id = ?', (id,)).fetchone()
        
        # Convert leave_date to date object if it's a string
        if isinstance(leave['date'], str):
            leave_date = datetime.strptime(leave['date'], '%Y-%m-%d').date()
        else:
            leave_date = leave['date']
        
        if action == 'approve':
            # Check for existing approved leave on same date (all types)
            existing = db.execute('''
                SELECT 1 FROM attendance 
                WHERE employee_id = ? 
                AND date = ?
                AND status = 'approved'
                AND session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
            ''', (leave['employee_id'], leave['date'])).fetchone()
            
            if existing:
                flash('Employee already has approved leave for this date', 'danger')
                return redirect(url_for('pending_leaves'))
            
            if leave['session_type'] == 'paid_leave':
                paid_leaves = get_leave_count(leave['employee_id'], leave_date.month, leave_date.year, db, 'paid_leave')
                if paid_leaves >= 1:
                    flash('Monthly paid leave limit reached for this employee', 'danger')
                    return redirect(url_for('pending_leaves'))
            
            elif leave['session_type'] in ['morning_half', 'afternoon_half', 'half_day']:
                half_days = get_leave_count(leave['employee_id'], leave_date.month, leave_date.year, db, 'half_day')
                if half_days >= 2:
                    flash('Monthly half-day limit reached for this employee', 'danger')
                    return redirect(url_for('pending_leaves'))
            
            elif leave['session_type'] == 'week_off':
                # Check if date is a holiday
                is_holiday = db.execute('SELECT 1 FROM holidays WHERE date = ?', 
                                      (leave['date'],)).fetchone()
                if is_holiday:
                    flash('Cannot approve week-off on a holiday', 'danger')
                    return redirect(url_for('pending_leaves'))
                
                # Calculate week boundaries (Monday to Sunday)
                week_start = leave_date - timedelta(days=leave_date.weekday())
                week_end = week_start + timedelta(days=6)
                
                # Check existing week-off in same week
                existing_week_off = db.execute('''
                    SELECT COUNT(*) FROM attendance 
                    WHERE employee_id = ? 
                    AND date BETWEEN ? AND ?
                    AND session_type = 'week_off'
                    AND status = 'approved'
                    AND id != ?
                ''', (leave['employee_id'], week_start, week_end, id)).fetchone()[0]
                
                if existing_week_off >= 1:
                    flash('Employee already has approved week-off this week', 'danger')
                    return redirect(url_for('pending_leaves'))
            
            db.execute('UPDATE attendance SET status = "approved" WHERE id = ?', (id,))
            flash('Leave approved successfully', 'success')
            
        elif action == 'reject':
            db.execute('UPDATE attendance SET status = "rejected" WHERE id = ?', (id,))
            flash('Leave rejected', 'info')
            
        db.commit()
    except Exception as e:
        logging.error(f"Leave handling error: {str(e)}", exc_info=True)
        flash('Error processing request', 'danger')
    
    return redirect(url_for('pending_leaves'))

@app.route('/api/employee/<int:id>/attendance')
@admin_required
def employee_attendance_api(id):
    try:
        db = get_db()
        records = db.execute('''
            SELECT date, session_type, status 
            FROM attendance 
            WHERE employee_id = ? 
            AND session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
            AND status IN ('approved', 'pending', 'rejected')
            ORDER BY date DESC
        ''', (id,)).fetchall()
        
        events = []
        for record in records:
            event_date = record['date']
            if isinstance(event_date, date):
                event_date = event_date.isoformat()
                
            # Check if this date is a holiday
            is_holiday = db.execute('SELECT 1 FROM holidays WHERE date = ?', 
                                  (record['date'],)).fetchone()
            
            # Skip if it's a holiday and the leave was cancelled
            if is_holiday and record['status'] == 'approved':
                continue
                
            events.append({
                'title': record['session_type'].replace('_', ' ').title(),
                'start': event_date,
                'allDay': True,
                'extendedProps': {
                    'session_type': record['session_type'],
                    'status': record['status']
                },
                'color': '#28a745' if record['session_type'] == 'paid_leave' else
                         '#dc3545' if record['session_type'] == 'loss_of_pay' else
                         '#ffc107' if record['session_type'] in ['half_day', 'morning_half', 'afternoon_half'] else
                         '#17a2b8' if record['session_type'] == 'week_off' else
                         '#6c757d'
            })
        return jsonify(events)
    except Exception as e:
        logging.error(f"Employee attendance API error: {str(e)}")
        return jsonify([])

@app.route('/api/employee/<int:id>/summary')
@admin_required
def employee_monthly_summary(id):
    month = request.args.get('month')
    year = request.args.get('year')
    
    if not month or not year:
        return jsonify({'error': 'Missing month or year'}), 400
    
    try:
        month = int(month)
        year = int(year)
    except ValueError:
        return jsonify({'error': 'Invalid month or year'}), 400
    
    db = get_db()
    
    working_days = db.execute('''
        SELECT COUNT(DISTINCT date) 
        FROM attendance 
        WHERE employee_id = ? 
        AND session_type = 'work'
        AND strftime('%w', date) != '0'
        AND strftime('%m', date) = ? 
        AND strftime('%Y', date) = ?
        AND NOT EXISTS (
            SELECT 1 FROM holidays 
            WHERE date = attendance.date
        )
        AND NOT EXISTS (
            SELECT 1 FROM attendance a2 
            WHERE a2.employee_id = attendance.employee_id
            AND a2.date = attendance.date
            AND a2.session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
            AND a2.status = 'approved'
        )
    ''', (id, f"{month:02d}", str(year))).fetchone()[0] or 0

    total_hours = db.execute('''
        SELECT SUM((
            strftime('%s', last_logout) - 
            strftime('%s', first_login)
          ))/3600.0
        FROM (
            SELECT 
                date,
                MIN(login_time) as first_login,
                MAX(logout_time) as last_logout
            FROM attendance
            WHERE employee_id = ?
            AND session_type = 'work'
            AND strftime('%m', date) = ?
            AND strftime('%Y', date) = ?
            GROUP BY date
        ) AS sub
    ''', (id, f"{month:02d}", str(year))).fetchone()[0] or 0

    paid_leaves = db.execute('''
        SELECT COUNT(*) FROM attendance 
        WHERE employee_id = ? 
        AND session_type = 'paid_leave'
        AND status = 'approved'
        AND strftime('%m', date) = ?
        AND strftime('%Y', date) = ?
    ''', (id, f"{month:02d}", str(year))).fetchone()[0] or 0

    loss_of_pay_leaves = db.execute('''
        SELECT COUNT(*) FROM attendance 
        WHERE employee_id = ? 
        AND session_type = 'loss_of_pay'
        AND status = 'approved'
        AND strftime('%m', date) = ?
        AND strftime('%Y', date) = ?
    ''', (id, f"{month:02d}", str(year))).fetchone()[0] or 0

    half_day_leaves = db.execute('''
        SELECT COUNT(*) FROM attendance 
        WHERE employee_id = ? 
        AND session_type IN ('half_day', 'morning_half', 'afternoon_half')
        AND status = 'approved'
        AND strftime('%m', date) = ?
        AND strftime('%Y', date) = ?
    ''', (id, f"{month:02d}", str(year))).fetchone()[0] or 0

    week_off_leaves = db.execute('''
        SELECT COUNT(*) FROM attendance 
        WHERE employee_id = ? 
        AND session_type = 'week_off'
        AND status = 'approved'
        AND strftime('%m', date) = ?
        AND strftime('%Y', date) = ?
    ''', (id, f"{month:02d}", str(year))).fetchone()[0] or 0


    holidays_count = db.execute('''
        SELECT COUNT(*) FROM holidays 
        WHERE strftime('%m', date) = ? 
        AND strftime('%Y', date) = ?
    ''', (f"{month:02d}", str(year))).fetchone()[0] or 0

    # return jsonify({
    #     'working_days': working_days,
    #     'total_hours': total_hours,
    #     'paid_leaves': paid_leaves,
    #     'loss_of_pay_leaves': loss_of_pay_leaves,
    #     'half_day_leaves': half_day_leaves,
    #     'week_off_leaves': week_off_leaves,
    #     'holidays': holidays_count  # Add this new field
    # })
    
    return jsonify({
        'working_days': working_days,
        'total_hours': total_hours,
        'paid_leaves': paid_leaves,
        'loss_of_pay_leaves': loss_of_pay_leaves,
        'half_day_leaves': half_day_leaves,
        'week_off_leaves': week_off_leaves,  # NEW
        'holidays': holidays_count
    })

def get_leave_count(employee_id, month, year, db, leave_type):
    """Get count of leaves for given type in specified month/year"""
    if leave_type == 'paid_leave':
        count = db.execute('''
            SELECT COUNT(*) FROM attendance
            WHERE employee_id = ?
            AND strftime('%m', date) = ?
            AND strftime('%Y', date) = ?
            AND session_type = 'paid_leave'
            AND status = 'approved'
        ''', (employee_id, f"{month:02d}", str(year))).fetchone()[0]
        return count
    
    elif leave_type == 'half_day':
        count = db.execute('''
            SELECT COUNT(*) FROM attendance
            WHERE employee_id = ?
            AND strftime('%m', date) = ?
            AND strftime('%Y', date) = ?
            AND session_type IN ('half_day', 'morning_half', 'afternoon_half')
            AND status = 'approved'
        ''', (employee_id, f"{month:02d}", str(year))).fetchone()[0]
        return count
    
    elif leave_type == 'week_off':
        count = db.execute('''
            SELECT COUNT(*) FROM attendance
            WHERE employee_id = ?
            AND strftime('%m', date) = ?
            AND strftime('%Y', date) = ?
            AND session_type = 'week_off'
            AND status = 'approved'
        ''', (employee_id, f"{month:02d}", str(year))).fetchone()[0]
        return count
    
    return 0

# Additional helper function for restoring leave counts
def restore_leave_count(employee_id, leave_type, month, year, db):
    """Restore leave count when a leave is cancelled due to holiday"""
    if leave_type == 'paid_leave':
        # Update paid leave count
        db.execute('''
            UPDATE employee_stats 
            SET paid_leaves = paid_leaves - 1 
            WHERE employee_id = ?
            AND month = ?
            AND year = ?
        ''', (employee_id, month, year))
    elif leave_type in ['morning_half', 'afternoon_half']:
        # Update half-day count
        db.execute('''
            UPDATE employee_stats 
            SET half_days = half_days - 1 
            WHERE employee_id = ?
            AND month = ?
            AND year = ?
        ''', (employee_id, month, year))
        
    elif leave_type == 'week_off':
        # Update week-off count
        db.execute('''
            UPDATE employee_stats 
            SET week_offs = week_offs - 1 
            WHERE employee_id = ?
            AND month = ?
            AND year = ?
        ''', (employee_id, month, year))

@app.route('/admin/holidays', methods=['GET', 'POST'])
@admin_required
def manage_holidays():
    if request.method == 'POST':
        try:
            date_str = request.form['date']
            description = request.form['description']
            holiday_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            db = get_db()
            
            # Check if holiday already exists
            existing = db.execute('SELECT 1 FROM holidays WHERE date = ?', 
                                 (holiday_date,)).fetchone()
            if existing:
                flash('Holiday already exists for this date', 'warning')
                return redirect(url_for('manage_holidays'))
            
            # Add new holiday
            db.execute('INSERT INTO holidays (date, description) VALUES (?, ?)',
                      (holiday_date, description))
            
            # Cancel existing leaves on this date
            leaves = db.execute('''
                SELECT id, employee_id, session_type, status 
                FROM attendance 
                WHERE date = ? 
                AND session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off', 'morning_half', 'afternoon_half')
                AND status = 'approved'
            ''', (holiday_date,)).fetchall()
            
            # Special handling for week-off leaves
            week_off_leaves = [leave for leave in leaves if leave['session_type'] == 'week_off']
            
            if week_off_leaves:
                # For each week-off leave, find a new available date in the same week
                for leave in week_off_leaves:
                    # Get the week boundaries (Monday to Sunday)
                    week_start = holiday_date - timedelta(days=holiday_date.weekday())
                    week_end = week_start + timedelta(days=6)
                    
                    # Find available dates in the same week (excluding holidays)
                    available_dates = []
                    current_date = week_start
                    while current_date <= week_end:
                        if current_date != holiday_date:
                            # Check if it's not already a holiday
                            is_holiday = db.execute('SELECT 1 FROM holidays WHERE date = ?', 
                                                  (current_date,)).fetchone()
                            if not is_holiday:
                                # Check if employee already has leave on this date
                                existing_leave = db.execute('''
                                    SELECT 1 FROM attendance 
                                    WHERE employee_id = ? 
                                    AND date = ?
                                    AND session_type IN ('paid_leave', 'loss_of_pay', 'half_day', 'week_off')
                                ''', (leave['employee_id'], current_date)).fetchone()
                                if not existing_leave:
                                    available_dates.append(current_date)
                        current_date += timedelta(days=1)
                    
                    if available_dates:
                        # Move the week-off to the first available date
                        new_date = available_dates[0]
                        db.execute('''
                            UPDATE attendance 
                            SET date = ?
                            WHERE id = ?
                        ''', (new_date, leave['id']))
                        flash(f"Moved week-off for employee {leave['employee_id']} to {new_date}", 'info')
                    else:
                        # No available dates, delete the week-off and restore leave count
                        db.execute('DELETE FROM attendance WHERE id = ?', (leave['id'],))
                        flash(f"No available dates to move week-off for employee {leave['employee_id']}", 'warning')
            
            # Delete other approved leaves (paid, loss of pay, half day)
            other_leaves = [leave for leave in leaves if leave['session_type'] != 'week_off']
            leave_ids = [leave['id'] for leave in leaves]
            if leave_ids:
                placeholders = ','.join(['?'] * len(leave_ids))
                db.execute(f'DELETE FROM attendance WHERE id IN ({placeholders})', leave_ids)
                flash(f'Deleted {len(leave_ids)} approved leave(s) on this holiday date', 'info')
            
            db.commit()
            
            flash(f'Holiday added successfully. Processed {len(leaves)} leave(s)', 'success')
        except Exception as e:
            logging.error(f"Error adding holiday: {str(e)}")
            flash('Failed to add holiday', 'danger')
    
    # For GET request, show the holidays
    db = get_db()
    holidays = db.execute('SELECT * FROM holidays ORDER BY date DESC').fetchall()
    return render_template('manage_holidays.html', 
                         holidays=holidays,
                         current_date=date.today())

@app.route('/admin/holiday/<int:id>/delete')
@admin_required
def delete_holiday(id):
    try:
        db = get_db()
        db.execute('DELETE FROM holidays WHERE id = ?', (id,))
        db.commit()
        flash('Holiday deleted successfully', 'success')
    except Exception as e:
        flash('Deletion failed', 'danger')
    
    return redirect(url_for('manage_holidays'))

@app.route('/api/holidays')
def get_holidays():
    try:
        db = get_db()
        holidays = db.execute('SELECT date, description FROM holidays').fetchall()
        events = []
        for holiday in holidays:
            events.append({
                'title': holiday['description'],
                'start': holiday['date'],
                'allDay': True,
                'color': "#ff69b4",  # Pink color
                'textColor': "#ffffff",
                'className': 'fc-event-holiday',
                'extendedProps': {
                    'isHoliday': True
                }
            })
        return jsonify(events)
    except Exception as e:
        logging.error(f"Error fetching holidays: {str(e)}")
        return jsonify([])
    
@app.route('/admin/employee/<int:id>/profile', methods=['GET', 'POST'])
@admin_required
def employee_profile(id):
    db = get_db()
    employee = db.execute('SELECT * FROM employees WHERE id = ?', (id,)).fetchone()
    profile = db.execute('SELECT * FROM employee_profiles WHERE employee_id = ?', (id,)).fetchone()

    if request.method == 'POST':
        # Handle form data
        employee_id = request.form.get('employee_id')
        dob = request.form.get('dob')
        
        # Handle file uploads
        photo = request.files.get('photo')
        id_card = request.files.get('id_card')
        
        photo_path = profile['photo_path'] if profile else None
        id_card_path = profile['id_card_path'] if profile else None
        
        # Process photo upload
        if photo and photo.filename and allowed_file(photo.filename, ALLOWED_IMAGE_EXTENSIONS):
            filename = secure_filename(f"photo_{id}_{photo.filename}")
            upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            photo.save(upload_path)
            photo_path = filename

        if id_card and id_card.filename and allowed_file(id_card.filename, ALLOWED_FILE_EXTENSIONS):
            filename = secure_filename(f"idcard_{id}_{id_card.filename}")
            upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            id_card.save(upload_path)
            id_card_path = filename
        
        # Update employee record with new ID
        db.execute('UPDATE employees SET employee_id = ? WHERE id = ?',
                  (employee_id, id))
        
        # Update or insert profile
        if profile:
            db.execute('''UPDATE employee_profiles 
                         SET photo_path = ?, id_card_path = ?, date_of_birth = ?
                         WHERE employee_id = ?''',
                     (photo_path, id_card_path, dob, id))
        else:
            db.execute('''INSERT INTO employee_profiles 
                        (employee_id, photo_path, id_card_path, date_of_birth)
                        VALUES (?, ?, ?, ?)''',
                     (id, photo_path, id_card_path, dob))
        
        db.commit()
        flash('Profile updated successfully', 'success')
        return redirect(url_for('employee_profile', id=id))
    
    return render_template('admin_employee_profile.html', 
                         employee=employee, 
                         profile=profile)

@app.route('/profile')
@employee_required
def employee_self_profile():
    db = get_db()
    employee = db.execute('SELECT * FROM employees WHERE id = ?', 
                         (session['employee_id'],)).fetchone()
    profile = db.execute('SELECT * FROM employee_profiles WHERE employee_id = ?', 
                        (session['employee_id'],)).fetchone()
    return render_template('employee_profile.html', 
                         employee=employee, 
                         profile=profile)
    
# app.py
@app.route('/admin/forgot_password', methods=['POST'])
def admin_forgot_password():
    try:
        email = request.form['email']
        db = get_db()
        admin = db.execute('SELECT * FROM admins WHERE email = ?', (email,)).fetchone()
        
        if not admin:
            flash('No admin account found with that email', 'danger')
            return redirect(url_for('admin_login'))
        
        # Generate random password
        new_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
        hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
        
        # Update database
        db.execute('UPDATE admins SET password = ? WHERE email = ?', 
                  (hashed_password.decode('utf-8'), email))
        db.commit()
        
        # Send email
        if send_password_reset_email(email, new_password):
            flash('Password reset email sent successfully', 'success')
        else:
            flash(f'Password reset to: {new_password} (email failed)', 'warning')
        
        return redirect(url_for('admin_login'))
            
    except Exception as e:
        logging.error(f"Password reset error: {str(e)}")
        flash('Failed to reset password', 'danger')
        return redirect(url_for('admin_login'))
    
    
@app.route('/admin/reset_password', methods=['POST'])
@admin_required
def admin_reset_password():
    if request.method == 'POST':
        try:
            current_password = request.form['current_password']
            new_password = request.form['new_password']
            confirm_password = request.form['confirm_password']
            
            if not current_password or not new_password or not confirm_password:
                flash('All fields are required', 'danger')
                return redirect(url_for('admin'))
            
            if new_password != confirm_password:
                flash('New passwords do not match', 'danger')
                return redirect(url_for('admin'))
                
            db = get_db()
            admin = db.execute('SELECT * FROM admins WHERE username = ?', 
                             (session['admin_username'],)).fetchone()
            
            if not bcrypt.checkpw(current_password.encode('utf-8'), admin['password'].encode('utf-8')):
                flash('Current password is incorrect', 'danger')
                return redirect(url_for('admin'))
                
            # Hash the new password
            hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
            
            db.execute('UPDATE admins SET password = ? WHERE username = ?',
                      (hashed_password.decode('utf-8'), session['admin_username']))
            db.commit()
            
            flash('Password updated successfully', 'success')
            return redirect(url_for('admin'))
            
        except Exception as e:
            logging.error(f"Password update error: {str(e)}")
            flash('Failed to update password', 'danger')
    
    return redirect(url_for('admin'))    

def send_password_reset_email(email, new_password):
    try:
        subject = "Your Admin Password Has Been Reset"
        body = f"New Password: {new_password}"
        
        msg = Message(
            subject,
            recipients=[email],
            body=body,
            sender=app.config['MAIL_DEFAULT_SENDER']
        )
        
        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"Email error: {str(e)}")
        raise e  # Re-raise to handle in calling function
    
    
# Teardown
@app.teardown_appcontext
def close_db(error):
    if hasattr(g, 'sqlite_db'):
        g.sqlite_db.close()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)



